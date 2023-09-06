import openpyxl
from openpyxl import styles
from openpyxl.styles import Font, Alignment


# abre el archivo de Excel existente
wb = openpyxl.load_workbook('estadisticas de glucosa.xlsx')

# selecciona la hoja de trabajo
hoja = wb.active

# combina las celdas A1 y B1 en las primeras tres filas
for i in range(1, 102):
    hoja.merge_cells(f'A{i}:E{i}')

import openpyxl

# Cargar archivo de Excel
workbook = openpyxl.load_workbook('estadisticas de glucosa.xlsx')

# Seleccionar hoja de Excel
worksheet = workbook['sheet1']

# Seleccionar rango de celdas que se desea centrar
cell_range = worksheet['A2:E101']

# Recorrer todas las celdas del rango y centrar su contenido
for row in cell_range:
    for cell in row:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

# Guardar archivo de Excel
workbook.save('ARCHIVOS XLSX/estadisticas de glucosa_centrado.xlsx')




